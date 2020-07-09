import openpyxl

class AmtReader(object):

    # Methoden
    # ------------------------------------
    def __init__(self, excel_file, sheet_name):
        '''
		Konstruktor
		'''
        self.excel_file = excel_file
        self.wb = openpyxl.load_workbook(self.excel_file)
        self.sheet_name = sheet_name
        self.sheet = self.wb[self.sheet_name]


    def get_columindex_by_name(self, column_name):
        liefereinheit_index = -99
        for row in self.sheet.iter_rows(min_row=1, max_row=1):
            for index, cell in enumerate(row):
                if cell.value == column_name:
                    liefereinheit_index = index
        
        return liefereinheit_index


    def get_oid_by_liefereinheit(self, liefereinheit):
        amt_oid = -99
        amt_oid_base = -99
        suffix = ""
        liefereinheit = unicode(liefereinheit)

        # NUPLWALD und NUPLKUEO haben nicht pro Liefereinheit einen Eintrag in der
        # zentralen AMT-Tabelle sondern nur einen einzigen. Deshalb muss das hier
        # erkannt werden
        if len(liefereinheit) == 5 and liefereinheit.endswith('03'):
            suffix = "." + liefereinheit[:3]
            liefereinheit = "9903"
        if len(liefereinheit) == 5 and liefereinheit.endswith('04'):
            suffix = "." + liefereinheit[:3]
            liefereinheit = "9904"

        amt_oid_index = self.get_columindex_by_name("AMT_OID")
        amt_name_de_index = self.get_columindex_by_name("AMT_NAME_DE")
        amt_name_fr_index = self.get_columindex_by_name("AMT_NAME_FR")
        amt_amtimweb_de_index = self.get_columindex_by_name("AMT_AMTIMWEB_DE")
        amt_amtimweb_fr_index = self.get_columindex_by_name("AMT_AMTIMWEB_FR")
        amt_liefereinheit_index = self.get_columindex_by_name("AMT_LIEFEREINHEIT")

        amt_oid = None
        amt_oid_base = None
        amt_name_de = None
        amt_name_fr = None
        amt_amtimweb_de = None
        amt_amtimweb_fr = None

        for row in self.sheet.iter_rows():
            amt_liefereinheit = unicode(row[amt_liefereinheit_index].value)
            if amt_liefereinheit == liefereinheit:
                amt_oid = row[amt_oid_index].value + suffix
                amt_oid_base = row[amt_oid_index].value
                amt_name_de = row[amt_name_de_index].value
                amt_name_fr = row[amt_name_fr_index].value
                amt_amtimweb_de = row[amt_amtimweb_de_index].value
                amt_amtimweb_fr = row[amt_amtimweb_fr_index].value

        return (amt_oid, amt_oid_base, amt_name_de, amt_name_fr, amt_amtimweb_de, amt_amtimweb_fr)