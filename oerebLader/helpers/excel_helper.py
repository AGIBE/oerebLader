import openpyxl

class AmtReader(object):

    # Methoden
    # ------------------------------------
    def __init__(self, excel_file, sheet_name):
        '''
		Konstruktor
		'''
        self.excel_file = excel_file
        self.wb = openpyxl.load_workbook(self.excel_file)
        self.sheet_name = sheet_name
        self.sheet = self.wb[self.sheet_name]


    def get_columindex_by_name(self, column_name):
        liefereinheit_index = -99
        for row in self.sheet.iter_rows(min_row=1, max_row=1):
            for index, cell in enumerate(row):
                if cell.value == column_name:
                    liefereinheit_index = index
        
        return liefereinheit_index


    def get_oid_by_liefereinheit(self, liefereinheit):
        amt_oid = -99
        amt_oid_base = -99
        suffix = ""
        liefereinheit = unicode(liefereinheit)

        # NUPLWALD und NUPLKUEO haben nicht pro Liefereinheit einen Eintrag in der
        # zentralen AMT-Tabelle sondern nur einen einzigen. Deshalb muss das hier
        # erkannt werden
        if len(liefereinheit) == 5 and liefereinheit.endswith('03'):
            suffix = "." + liefereinheit[:3]
            liefereinheit = "9903"
        if len(liefereinheit) == 5 and liefereinheit.endswith('04'):
            suffix = "." + liefereinheit[:3]
            liefereinheit = "9904"

        amt_oid_index = self.get_columindex_by_name("AMT_OID")
        amt_liefereinheit_index = self.get_columindex_by_name("AMT_LIEFEREINHEIT")

        for row in self.sheet.iter_rows():
            amt_liefereinheit = unicode(row[amt_liefereinheit_index].value)
            if amt_liefereinheit == liefereinheit:
                amt_oid = row[amt_oid_index].value + suffix
                amt_oid_base = row[amt_oid_index].value

        return (amt_oid, amt_oid_base)

if __name__ == "__main__":
    ar = AmtReader(r"\\geoda.infra.be.ch\arbeit\Anwend\OEREBK\Data\AMT\AMT.xlsx","AMT")
    oids = ar.get_oid_by_liefereinheit(36204)
    print(oids)